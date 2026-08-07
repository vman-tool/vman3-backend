"""
Enrich the VA question dictionary from an uploaded XLSForm (xForm) workbook.

ODK's question sync (odk/services/data_download.fetch_form_questions) tells us
which questions exist, but the labels and choice lists it returns are often
thin. The XLSForm that was used to deploy the instrument carries the full
`survey` and `choices` sheets, so uploading it fills those gaps.

This is deliberately non-destructive:
  - questions are matched to existing records by `name`
  - a field is only written when the stored value is missing or blank
  - names not already present are reported as `unmatched`, never inserted,
    so ODK sync remains the source of truth for *which* questions exist

That makes the upload safe to re-run, and means a wrong workbook cannot wipe
the dictionary.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from arango.database import StandardDatabase
from fastapi.concurrency import run_in_threadpool
from openpyxl import load_workbook

from app.odk.models.questions_models import VA_Question
from app.shared.configs.models import ResponseMainModel
from app.shared.middlewares.exceptions import BadRequestException


def _clean(value: Any) -> Optional[str]:
    """Trim a cell to a non-empty string, or None."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_blank(value: Any) -> bool:
    """True when a stored field carries no usable content."""
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    if isinstance(value, (list, tuple, dict)):
        return len(value) == 0
    return False


def _header_index(row: Tuple) -> Dict[str, int]:
    """Map normalised header names to column positions.

    XLSForm label/hint columns are language-tagged - `label::English(en)`,
    `label::Swahili (sw)` - so match on the prefix before '::' and take the
    first one seen, which is the workbook's primary language.
    """
    index: Dict[str, int] = {}
    for pos, cell in enumerate(row):
        name = _clean(cell)
        if not name:
            continue
        key = name.split("::")[0].strip().lower()
        index.setdefault(key, pos)
    return index


def _cell(row: Tuple, index: Dict[str, int], key: str) -> Optional[str]:
    pos = index.get(key)
    if pos is None or pos >= len(row):
        return None
    return _clean(row[pos])


def _label_columns(row: Tuple) -> Dict[str, int]:
    """Map each label language to its column position.

    XLSForm tags label columns by language, and the spacing varies between
    authors:  `label::English(en)`, `label::Siswati (es)`. An untagged
    `label` column is recorded as 'Default'.

    Returns e.g. {'English': 3, 'Swahili': 4}. The language name (not the
    code) is used as the key, since that is what gets shown in the UI.
    """
    languages: Dict[str, int] = {}
    for pos, cell in enumerate(row):
        name = _clean(cell)
        if not name or not name.lower().startswith("label"):
            continue
        if "::" not in name:
            languages.setdefault("Default", pos)
            continue
        tag = name.split("::", 1)[1].strip()
        # strip the trailing '(code)' -> 'English (en)' becomes 'English'
        language = tag.split("(")[0].strip() or tag
        languages.setdefault(language, pos)
    return languages


def _choice_list_name(type_value: Optional[str]) -> Optional[str]:
    """Extract the choice-list name from an XLSForm `type` cell.

    'select_one yes_no'      -> 'yes_no'
    'select multiple sympt'  -> 'sympt'
    'select_one select_2'    -> 'select_2'

    Both the underscored and spaced spellings of the keyword occur in the wild.
    Only the keyword is normalised - the list name is returned verbatim, since
    underscores are significant in it (an earlier version stripped them and
    silently failed to match lists like `select_2`).
    """
    text = (type_value or "").strip()
    if not text:
        return None
    lowered = text.lower()
    for prefix in ("select_one ", "select one ", "select_multiple ", "select multiple "):
        if lowered.startswith(prefix):
            remainder = text[len(prefix):].strip()
            # trailing modifiers such as `or_other` may follow the list name
            return remainder.split()[0] if remainder else None
    return None


def _merge_option_labels(
    stored: List[Dict],
    incoming: List[Dict],
    override_labels: bool,
) -> List[Dict]:
    """Fold the workbook's per-language answer text into stored options.

    Matched by option `value`, since that is the only stable identifier - the
    ordering of a choice list is not guaranteed to survive an ODK round trip.
    Options with no counterpart in the workbook are passed through untouched.
    """
    by_value = {str(o.get("value")): o for o in incoming if o.get("value") is not None}

    merged: List[Dict] = []
    for option in stored:
        option = dict(option)
        source = by_value.get(str(option.get("value")))
        if source and source.get("labels"):
            existing = dict(option.get("labels") or {})
            if override_labels:
                option["labels"] = {**existing, **source["labels"]}
            else:
                option["labels"] = {**existing, **{
                    lang: text
                    for lang, text in source["labels"].items()
                    if lang not in existing
                }}
        merged.append(option)
    return merged


def parse_xform(content: bytes) -> Tuple[Dict[str, Dict], Dict[str, List[Dict]]]:
    """Read the survey and choices sheets out of an XLSForm workbook.

    Returns (questions_by_name, choices_by_list_name). Raises BadRequestException
    when the workbook is not a recognisable XLSForm.
    """
    from io import BytesIO

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise BadRequestException(f"Could not read the workbook: {exc}")

    if "survey" not in workbook.sheetnames:
        raise BadRequestException(
            "This does not look like an XLSForm - no 'survey' sheet was found. "
            f"Sheets present: {', '.join(workbook.sheetnames) or 'none'}."
        )

    # --- choices sheet: list_name -> [{path,value,label,labels}] -------------
    # The choices sheet is language-tagged exactly like the survey sheet, so the
    # answer text is collected per language too. Without this a viewer switched
    # to Swahili would show Swahili questions against English answers.
    choices: Dict[str, List[Dict]] = {}
    if "choices" in workbook.sheetnames:
        rows = workbook["choices"].iter_rows(values_only=True)
        header_row = next(rows, ()) or ()
        header = _header_index(header_row)
        choice_label_cols = _label_columns(header_row)
        for row in rows:
            list_name = _cell(row, header, "list_name") or _cell(row, header, "list name")
            value = _cell(row, header, "name")
            if not list_name or not value:
                continue
            labels = {
                language: _clean(row[pos])
                for language, pos in choice_label_cols.items()
                if pos < len(row) and _clean(row[pos])
            }
            primary = labels.get("English") or next(iter(labels.values()), None)
            choices.setdefault(list_name.lower(), []).append(
                {
                    "path": "",
                    "value": value,
                    "label": primary or value,
                    "labels": labels or None,
                }
            )

    # --- survey sheet: name -> {type, labels:{lang:text}, choice_list, path} -
    questions: Dict[str, Dict] = {}
    rows = workbook["survey"].iter_rows(values_only=True)
    header_row = next(rows, ()) or ()
    header = _header_index(header_row)
    label_cols = _label_columns(header_row)

    # Track begin/end group|repeat so each question gets an ODK-style path
    # (/consented/deceased_crvs/.../name) rather than a bare name.
    group_stack: List[str] = []

    for row in rows:
        name = _cell(row, header, "name")
        type_value = _cell(row, header, "type")
        kind = (type_value or "").strip().lower()

        if kind.startswith(("begin group", "begin_group", "begin repeat", "begin_repeat")):
            if name:
                group_stack.append(name)
            continue
        if kind.startswith(("end group", "end_group", "end repeat", "end_repeat")):
            if group_stack:
                group_stack.pop()
            continue
        if not name:
            continue

        labels = {
            language: _clean(row[pos])
            for language, pos in label_cols.items()
            if pos < len(row) and _clean(row[pos])
        }
        questions[name] = {
            "type": type_value,
            "labels": labels,
            "choice_list": _choice_list_name(type_value),
            "path": "/" + "/".join(group_stack + [name]),
        }

    if not questions:
        raise BadRequestException("The 'survey' sheet contained no named questions.")

    return questions, choices


# The WHO 2022 VA instrument, shipped with the application so that every
# deployment has an English data dictionary before ODK is ever contacted.
# It lives inside the `app` package (not backend/resources) for two reasons:
# the Dockerfile only does `COPY app ./app/`, and resolving from
# Path(__file__) makes it independent of the working directory.
DEFAULT_XFORM = (
    Path(__file__).resolve().parents[2] / "resources" / "whova2022_xls_form_for_odk.xlsx"
)

# Structural XLSForm rows that describe layout or computation rather than a
# variable the interviewer answers.
_NON_QUESTION_TYPES = {"note", "calculate", "start", "end", "today", "deviceid"}


async def seed_dictionary_if_empty(db: StandardDatabase) -> int:
    """Populate the dictionary from the bundled WHO form when it is empty.

    Runs at startup. If any questions already exist this does nothing at all -
    an established dictionary is never touched, and ODK sync remains the
    authority on which questions a given deployment actually has.

    Returns the number of questions created (0 when it was a no-op).
    """
    try:
        existing = await VA_Question.get_many(paging=False, db=db)
        if existing:
            return 0

        if not DEFAULT_XFORM.exists():
            print(f"Default xForm not found at {DEFAULT_XFORM}; skipping dictionary seed")
            return 0

        parsed, choices = await run_in_threadpool(parse_xform, DEFAULT_XFORM.read_bytes())

        created = 0
        for name, source in parsed.items():
            kind = (source.get("type") or "").strip().lower()
            if kind in _NON_QUESTION_TYPES:
                continue

            labels = source.get("labels") or {}
            label = labels.get("English") or (next(iter(labels.values()), None) or name)
            options = None
            if source.get("choice_list"):
                options = choices.get(source["choice_list"].lower())

            await VA_Question(
                # names are stored lower-case, matching ODK's normalisation
                name=name.lower(),
                path=source.get("path") or f"/{name}",
                type=source.get("type") or "string",
                label=label,
                labels=labels or None,
                options=options,
            ).save(db, "name")
            created += 1

        print(f"Seeded data dictionary with {created} questions from {DEFAULT_XFORM.name}")
        return created
    except Exception as exc:  # seeding must never block application startup
        print(f"Dictionary seeding skipped: {exc}")
        return 0


async def get_data_dictionary(
    db: StandardDatabase,
    include_options: bool = False,
) -> ResponseMainModel:
    """Return the stored VA question dictionary for display.

    Read-only: unlike POST /odk/fetch_form_questions this does not contact ODK
    Central, it just reports what is currently stored. `languages` lists every
    language present across the dictionary so the UI can render one column per
    language without scanning the rows itself.

    :param include_options: also return each question's choice list. Off by
        default because the settings table only needs the count, and the full
        lists roughly quadruple the payload. The VA record viewer asks for
        them so it can translate answers as well as questions.
    """
    stored = await VA_Question.get_many(paging=False, db=db)

    rows = []
    languages: set = set()
    for record in stored:
        labels = record.get("labels") or {}
        languages.update(labels.keys())
        options = record.get("options") or []
        row = {
            "name": record.get("name"),
            "label": record.get("label"),
            "labels": labels,
            "type": record.get("type"),
            "options_count": len(options),
        }
        if include_options:
            row["options"] = options
        rows.append(row)

    rows.sort(key=lambda r: (r.get("name") or "").lower())

    # English first when present, remaining languages alphabetically
    ordered = sorted(languages, key=lambda l: (l != "English", l.lower()))

    return ResponseMainModel(
        data={"questions": rows, "languages": ordered},
        message=f"{len(rows)} questions in the data dictionary",
        total=len(rows),
    )


async def update_question_label(
    name: str,
    language: str,
    label: str,
    db: StandardDatabase,
) -> ResponseMainModel:
    """Edit one question's label for one language.

    Used by the inline editor in the data dictionary table, mainly to tidy up
    the WHO instrument's embedded ids - "(Id10010) [Name of VA interviewer]"
    reads better as "Name of VA interviewer".

    Only the given language is touched. When it is the primary language the
    legacy top-level `label` is kept in step, since consumers that predate
    multi-language support still read that field.
    """
    cleaned = (label or "").strip()
    if not cleaned:
        raise BadRequestException("The label cannot be empty.")
    if not (language or "").strip():
        raise BadRequestException("A language must be given.")

    matches = await VA_Question.get_many(filters={"name": name}, db=db)
    if not matches:
        raise BadRequestException(f"No question named '{name}' exists in the dictionary.")

    record = matches[0]
    labels = dict(record.get("labels") or {})
    labels[language] = cleaned

    doc = {k: v for k, v in record.items() if not k.startswith("_")}
    doc["labels"] = labels

    # `label` mirrors the primary language: English when the question has it,
    # otherwise whichever language is being edited on a single-language record.
    primary = "English" if "English" in labels else next(iter(labels), None)
    if language == primary:
        doc["label"] = cleaned

    await VA_Question(**doc).save(db, "name")

    return ResponseMainModel(
        data={"name": name, "language": language, "label": cleaned, "labels": labels},
        message=f"Updated the {language} label for {name}.",
    )


async def enrich_questions_from_xform(
    content: bytes,
    filename: str,
    db: StandardDatabase,
    override_labels: bool = False,
) -> ResponseMainModel:
    """Parse an uploaded XLSForm and fill gaps in the stored VA questions.

    :param override_labels: when False (the default) an existing label for a
        given language is kept and only new languages are added. When True the
        workbook's text replaces what is stored for the languages it carries -
        languages it does not carry are still left untouched.
    """
    parsed, choices = await run_in_threadpool(parse_xform, content)

    stored = await VA_Question.get_many(paging=False, db=db)
    if not stored:
        raise BadRequestException(
            "There are no synchronized questions to enrich. Run Server "
            "Synchronization first, then upload the xForm."
        )

    updated: List[str] = []
    filled = {"labels": 0, "label": 0, "type": 0, "options": 0}

    # Match case-insensitively: ODK normalises question names to lower case
    # (`id10010b`) while XLSForm preserves the author's casing (`Id10010b`).
    # Exact matching resolved only 19 of 521 names on a real WHO instrument;
    # case-insensitive matching resolves 455.
    by_lower = {name.lower(): value for name, value in parsed.items()}

    for record in stored:
        name = record.get("name") or ""
        source = by_lower.get(name.lower())
        if not source:
            continue

        changes: Dict[str, Any] = {}
        incoming = source.get("labels") or {}

        # Languages merge additively: an existing language is never overwritten,
        # new ones are added. This is what lets a Swahili xForm and an Eswatini
        # xForm both contribute to the same question over successive uploads.
        if incoming:
            existing = dict(record.get("labels") or {})
            if override_labels:
                # workbook wins for the languages it carries; others are kept
                merged_labels = {**existing, **incoming}
            else:
                merged_labels = {**existing, **{
                    lang: text for lang, text in incoming.items() if lang not in existing
                }}
            if merged_labels != existing:
                changes["labels"] = merged_labels

        # `label` stays the primary display value for consumers that predate
        # multi-language support. Filled when blank, or always when overriding.
        if incoming and (override_labels or _is_blank(record.get("label"))):
            preferred = incoming.get("English") or next(iter(incoming.values()))
            if preferred and preferred != record.get("label"):
                changes["label"] = preferred

        if _is_blank(record.get("type")) and source.get("type"):
            changes["type"] = source["type"]

        if source.get("choice_list"):
            incoming_options = choices.get(source["choice_list"].lower())
            if incoming_options:
                if _is_blank(record.get("options")):
                    changes["options"] = incoming_options
                else:
                    # Options are already stored (usually from ODK, single
                    # language). Merge the workbook's per-language answer text
                    # into them by value, using the same additive rule as the
                    # question labels.
                    merged_options = _merge_option_labels(
                        record.get("options") or [], incoming_options, override_labels
                    )
                    if merged_options != (record.get("options") or []):
                        changes["options"] = merged_options

        if not changes:
            continue

        merged = {**record, **changes}
        # get_many returns raw documents; drop Arango bookkeeping keys before
        # re-validating through the model.
        merged.pop("_rev", None)
        await VA_Question(**{k: v for k, v in merged.items() if not k.startswith("_")}).save(db, "name")

        updated.append(record["name"])
        for field in changes:
            filled[field] += 1

    languages_seen = sorted({
        lang for q in parsed.values() for lang in (q.get("labels") or {})
    })
    stored_names = {(r.get("name") or "").lower() for r in stored}
    unmatched = sorted(n for n in parsed if n.lower() not in stored_names)

    return ResponseMainModel(
        data={
            "file": filename,
            "questions_in_xform": len(parsed),
            "choice_lists_in_xform": len(choices),
            "questions_stored": len(stored),
            "questions_updated": len(updated),
            "fields_filled": filled,
            "languages_seen": languages_seen,
            "unmatched_in_xform": unmatched[:50],
            "unmatched_count": len(unmatched),
        },
        message=(
            f"Enriched {len(updated)} of {len(stored)} questions from {filename}."
            if updated
            else "No gaps were found - the stored questions already have labels, types and options."
        ),
        total=len(updated),
    )
