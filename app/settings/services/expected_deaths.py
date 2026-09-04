"""Expected number of deaths per administrative unit, sourced from an xForm's
`choices` sheet.

The hierarchy (how many levels, what they are called) is never assumed -
country deployments differ (Region/District/Ward for Tanzania is just one
example). It is derived purely from the workbook's own data: every choice
list row may carry a `parent` value naming another row's `name` - wherever
that points resolves which choice list is this one's parent, exactly as the
xForm itself expresses the relationship. A workbook with none of its columns
starting with `expected_deaths` has nothing to import - callers get `None`
back and skip silently.

Values are per period: one or more `expected_deaths::<period>` columns (e.g.
`expected_deaths::2023`, `expected_deaths::2024`), or a single bare
`expected_deaths` column treated as the one period "total". A node may carry
its own raw value directly, or have none and only ever be a computed sum of
its children - both are valid at any level, so a district with a full ward
breakdown and a district with only a single top-line figure are both handled
the same way. Aggregates are recomputed, per period, on every import and
every edit so totals can never drift out of sync.

A row with no value in any period, and no descendant that ends up with one
either, is not stored at all - this is what keeps placeholder choices like
"other"/"dk" out of the tree. Should a later upload give such a row a real
value, it is created then, not before.
"""

import hashlib
from io import BytesIO
from typing import Dict, List, Optional

from arango.database import StandardDatabase
from fastapi.concurrency import run_in_threadpool
from openpyxl import load_workbook

from app.settings.services.xform_dictionary import _cell, _clean, _header_index, _label_columns
from app.shared.configs.constants import db_collections
from app.shared.configs.models import ResponseMainModel
from app.shared.middlewares.exceptions import BadRequestException


def _to_int(value) -> Optional[int]:
    """Coerce an Excel cell to a whole number (decimals dropped), or None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return round(float(text))
    except ValueError:
        return None


def _node_key(parent_key: Optional[str], list_name: str, value: str) -> str:
    """A node's identity is its full path (parent chain + own value), not just
    its own (list_name, value) - `name` is only unique within a list_name for
    the "real" choices. Placeholder options like "other"/"dk" are commonly
    repeated verbatim under every parent in a list, so (list_name, value)
    alone would collapse every "other" district in the country into one
    node, silently dropping the rest.
    """
    return hashlib.sha1(f"{parent_key or ''}|{list_name}:{value}".encode("utf-8")).hexdigest()


def _period_columns(header_row) -> Dict[str, int]:
    """Map each `expected_deaths::<period>` column to its position.

    A bare `expected_deaths` column (no `::`) is treated as the single
    period named "total". Mirrors how label::<language> columns work.
    """
    periods: Dict[str, int] = {}
    for pos, cell in enumerate(header_row):
        name = _clean(cell)
        if not name or not name.lower().startswith("expected_deaths"):
            continue
        if "::" not in name:
            periods.setdefault("total", pos)
            continue
        period = name.split("::", 1)[1].strip()
        if period:
            periods.setdefault(period, pos)
    return periods


def parse_expected_deaths_hierarchy(content: bytes) -> Optional[List[Dict]]:
    """Read the `choices` sheet and build the admin-unit hierarchy.

    Returns a flat list of node dicts (`_key`, `level`, `list_name`, `value`,
    `label`, `parent_key`, `raw_expected_deaths`), ordered top level first -
    or `None` when the workbook has no `expected_deaths` column(s), or isn't
    a readable XLSForm at all. Rows with no value in any period and no
    surviving descendant are already excluded.
    """
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return None

    if "choices" not in workbook.sheetnames:
        return None

    rows = workbook["choices"].iter_rows(values_only=True)
    header_row = next(rows, ()) or ()
    header = _header_index(header_row)

    period_columns = _period_columns(header_row)
    if not period_columns:
        return None

    list_name_pos = header.get("list_name", header.get("list name"))
    name_pos = header.get("name")
    if list_name_pos is None or name_pos is None:
        return None

    label_cols = _label_columns(header_row)

    raw_rows: Dict[str, List[Dict]] = {}
    for row in rows:
        list_name = _cell(row, header, "list_name") or _cell(row, header, "list name")
        value = _cell(row, header, "name")
        if not list_name or not value:
            continue
        list_name = list_name.strip().lower()

        labels = {
            language: _clean(row[pos])
            for language, pos in label_cols.items()
            if pos < len(row) and _clean(row[pos])
        }
        label = labels.get("English") or next(iter(labels.values()), None) or value

        parent_value = _cell(row, header, "parent")

        deaths_by_period: Dict[str, int] = {}
        for period, pos in period_columns.items():
            if pos < len(row):
                number = _to_int(row[pos])
                if number is not None:
                    deaths_by_period[period] = number

        raw_rows.setdefault(list_name, []).append({
            "value": value,
            "label": label,
            "parent_value": parent_value,
            "deaths_by_period": deaths_by_period,
        })

    if not raw_rows:
        return None

    # For each choice list, resolve its parent list by looking up which
    # other list each of its rows' `parent` value belongs to (a `name` is
    # only guaranteed unique within its own list, so this is a majority
    # vote across the list's own rows, not a single lookup).
    name_to_list_names: Dict[str, set] = {}
    for list_name, entries in raw_rows.items():
        for entry in entries:
            name_to_list_names.setdefault(entry["value"], set()).add(list_name)

    parent_list_name: Dict[str, Optional[str]] = {}
    for list_name, entries in raw_rows.items():
        votes: Dict[str, int] = {}
        for entry in entries:
            parent_value = entry["parent_value"]
            if not parent_value:
                continue
            for candidate in name_to_list_names.get(parent_value, ()):
                if candidate != list_name:
                    votes[candidate] = votes.get(candidate, 0) + 1
        parent_list_name[list_name] = max(votes, key=votes.get) if votes else None

    # The relevant hierarchy is whichever chain of lists actually carries
    # period values somewhere in it. Only the first such list found is used -
    # a workbook with more than one unrelated expected_deaths hierarchy is
    # not a case this supports.
    data_bearing = [ln for ln, entries in raw_rows.items() if any(e["deaths_by_period"] for e in entries)]
    if not data_bearing:
        return None

    chain: List[str] = []
    current: Optional[str] = data_bearing[0]
    seen = set()
    while current and current not in seen:
        chain.append(current)
        seen.add(current)
        current = parent_list_name.get(current)
    chain.reverse()

    nodes: List[Dict] = []
    key_by_value_at_level: Dict[str, Dict[str, str]] = {}

    for level_index, list_name in enumerate(chain, start=1):
        parent_list = chain[level_index - 2] if level_index > 1 else None
        key_by_value_at_level[list_name] = {}
        for entry in raw_rows.get(list_name, []):
            parent_key = None
            if parent_list:
                parent_value = entry["parent_value"]
                if parent_value:
                    # Looked up by value within the immediate parent level -
                    # the one link the file itself expresses. If that parent
                    # level has its own same-named siblings (e.g. two
                    # "other" placeholders), this resolves to whichever one
                    # was processed last; a real, non-placeholder duplicate
                    # name at the same level is an ambiguity in the source
                    # file itself, not something resolvable here.
                    parent_key = key_by_value_at_level.get(parent_list, {}).get(parent_value)
            key = _node_key(parent_key, list_name, entry["value"])
            nodes.append({
                "_key": key,
                "level": level_index,
                "list_name": list_name,
                "value": entry["value"],
                "label": entry["label"],
                "parent_key": parent_key,
                "raw_expected_deaths": entry["deaths_by_period"],
            })
            key_by_value_at_level[list_name][entry["value"]] = key

    return _drop_empty_branches(nodes)


def _drop_empty_branches(nodes: List[Dict]) -> List[Dict]:
    """Remove rows with no value in any period and no surviving descendant -
    e.g. an "other"/"dk" placeholder with nothing under it. A node with no
    raw value of its own is kept as long as at least one descendant survives,
    since it is still a meaningful container for those.
    """
    if not nodes:
        return nodes

    children_of: Dict[Optional[str], List[str]] = {}
    for n in nodes:
        children_of.setdefault(n["parent_key"], []).append(n["_key"])

    keep: set = set()
    max_level = max(n["level"] for n in nodes)
    for level in range(max_level, 0, -1):
        for n in nodes:
            if n["level"] != level:
                continue
            has_children = any(child_key in keep for child_key in children_of.get(n["_key"], []))
            if n["raw_expected_deaths"] or has_children:
                keep.add(n["_key"])

    return [n for n in nodes if n["_key"] in keep]


def recompute_aggregates(by_key: Dict[str, Dict]) -> None:
    """Compute each node's displayed expected_deaths, per period, in place.

    A node's own raw value for a period wins when present - a unit with both
    a top-line total and a partial per-child breakdown is not silently
    replaced by a possibly-incomplete sum. Otherwise the period is the sum of
    whatever children have a value for it (children with nothing for that
    period are simply not counted), or absent entirely if no child has it
    either.

    Bottom-up by level, so a level's totals are only ever built from
    already-finalised children.
    """
    if not by_key:
        return

    children: Dict[Optional[str], List[Dict]] = {}
    for doc in by_key.values():
        children.setdefault(doc.get("parent_key"), []).append(doc)

    max_level = max(doc["level"] for doc in by_key.values())
    for level in range(max_level, 0, -1):
        for doc in by_key.values():
            if doc["level"] != level:
                continue
            raw = doc.get("raw_expected_deaths") or {}
            kids = children.get(doc["_key"], [])

            periods = set(raw.keys())
            for kid in kids:
                periods.update((kid.get("expected_deaths") or {}).keys())

            computed: Dict[str, int] = {}
            for period in periods:
                if period in raw:
                    computed[period] = raw[period]
                else:
                    values = [
                        kid["expected_deaths"][period]
                        for kid in kids
                        if period in (kid.get("expected_deaths") or {})
                    ]
                    if values:
                        computed[period] = sum(values)
            doc["expected_deaths"] = computed


def _build_tree(by_key: Dict[str, Dict]) -> List[Dict]:
    children: Dict[Optional[str], List[Dict]] = {}
    for doc in by_key.values():
        children.setdefault(doc.get("parent_key"), []).append(doc)
    for kids in children.values():
        kids.sort(key=lambda d: (d.get("label") or "").lower())

    def build(doc: Dict) -> Dict:
        kids = children.get(doc["_key"], [])
        return {
            "key": doc["_key"],
            "level": doc["level"],
            "value": doc["value"],
            "label": doc["label"],
            "expected_deaths": doc.get("expected_deaths") or {},
            "is_leaf": len(kids) == 0,
            "children": [build(kid) for kid in kids],
        }

    roots = children.get(None, [])
    return [build(root) for root in roots]


def _periods_present(by_key: Dict[str, Dict]) -> List[str]:
    periods: set = set()
    for doc in by_key.values():
        periods.update((doc.get("expected_deaths") or {}).keys())
    # "total" (the bare-column, single-period case) first if present,
    # otherwise sorted - which sorts years ascending as a side effect.
    return sorted(periods, key=lambda p: (p != "total", p))


async def _fetch_all(db: StandardDatabase) -> Dict[str, Dict]:
    if not db.has_collection(db_collections.EXPECTED_DEATHS):
        return {}

    def execute():
        cursor = db.aql.execute(f"FOR doc IN {db_collections.EXPECTED_DEATHS} RETURN doc")
        return {doc["_key"]: doc for doc in cursor}

    return await run_in_threadpool(execute)


async def _write_all(db: StandardDatabase, docs: List[Dict]) -> None:
    if not docs:
        return

    def execute():
        if not db.has_collection(db_collections.EXPECTED_DEATHS):
            db.create_collection(db_collections.EXPECTED_DEATHS)
        collection = db.collection(db_collections.EXPECTED_DEATHS)
        collection.insert_many(docs, overwrite=True, overwrite_mode="update")

    await run_in_threadpool(execute)


async def import_expected_deaths_from_xform(content: bytes, db: StandardDatabase) -> Optional[Dict]:
    """Import (or refresh) the admin-unit hierarchy from an uploaded xForm.

    Non-destructive like the rest of the dictionary upload, per period: a
    period already stored for a node is kept as-is (it may have been edited
    since) - only a period genuinely new to that node is taken from the
    file. Labels and structure always refresh, since those are not
    user-editable. A row absent from the file (pruned as empty, or removed
    entirely) is never deleted - it simply is not touched.

    Returns a summary dict, or None when the workbook carries no
    `expected_deaths` column(s) at all - the caller should skip silently.
    """
    parsed_nodes = await run_in_threadpool(parse_expected_deaths_hierarchy, content)
    if parsed_nodes is None:
        return None

    existing = await _fetch_all(db)

    max_level = max(node["level"] for node in parsed_nodes) if parsed_nodes else 0
    created = 0

    for node in parsed_nodes:
        prior = existing.get(node["_key"])
        doc = dict(node)
        raw = dict(node.get("raw_expected_deaths") or {})

        if prior is None:
            created += 1
        else:
            prior_raw = prior.get("raw_expected_deaths") or {}
            for period, value in prior_raw.items():
                if period in raw:
                    raw[period] = value

        doc["raw_expected_deaths"] = raw
        existing[node["_key"]] = doc

    recompute_aggregates(existing)
    await _write_all(db, list(existing.values()))

    return {
        "levels": max_level,
        "nodes_in_file": len(parsed_nodes),
        "nodes_created": created,
        "periods": _periods_present(existing),
    }


async def get_expected_deaths_tree(db: StandardDatabase) -> ResponseMainModel:
    existing = await _fetch_all(db)
    if not existing:
        return ResponseMainModel(
            data={"configured": False, "max_level": 0, "periods": [], "tree": []},
            message="No expected deaths data has been imported yet.",
        )

    max_level = max(doc["level"] for doc in existing.values())
    return ResponseMainModel(
        data={
            "configured": True,
            "max_level": max_level,
            "periods": _periods_present(existing),
            "tree": _build_tree(existing),
        },
        message="Expected deaths fetched successfully",
    )


async def update_expected_deaths_value(key: str, period: str, expected_deaths: int, db: StandardDatabase) -> ResponseMainModel:
    if expected_deaths < 0:
        raise BadRequestException("Expected deaths cannot be negative.")
    if not (period or "").strip():
        raise BadRequestException("A period is required.")

    existing = await _fetch_all(db)
    doc = existing.get(key)
    if not doc:
        raise BadRequestException("Unknown administrative unit.")

    children = [d for d in existing.values() if d.get("parent_key") == key]
    if children:
        raise BadRequestException(
            "Only an administrative unit with no children can be edited directly - totals above it "
            "are calculated automatically from their children."
        )

    raw = dict(doc.get("raw_expected_deaths") or {})
    raw[period] = round(expected_deaths)
    doc["raw_expected_deaths"] = raw

    recompute_aggregates(existing)
    await _write_all(db, list(existing.values()))

    max_level = max(d["level"] for d in existing.values())
    return ResponseMainModel(
        data={
            "configured": True,
            "max_level": max_level,
            "periods": _periods_present(existing),
            "tree": _build_tree(existing),
        },
        message=f"Updated expected deaths for {doc.get('label')}.",
    )
