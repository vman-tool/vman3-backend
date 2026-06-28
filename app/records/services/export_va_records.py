import asyncio
from datetime import date
from io import BytesIO
from typing import List, Optional

import pandas as pd
from arango.database import StandardDatabase
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from fastapi.concurrency import run_in_threadpool

from app.settings.services.odk_configs import fetch_odk_config
from app.shared.configs.constants import db_collections
from app.shared.configs.security import get_location_limit_values


def _strip_workbook_formatting(workbook) -> None:
    """Remove all bold, borders, and fills from every cell in the workbook."""
    from openpyxl.styles import Font, Border, PatternFill, Alignment
    plain_font = Font(name='Calibri', size=11, bold=False)
    no_border  = Border()
    no_fill    = PatternFill(fill_type=None)
    no_align   = Alignment()
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font      = plain_font
                cell.border    = no_border
                cell.fill      = no_fill
                cell.alignment = no_align


async def _run_export_query(query: str, bind_vars: dict, db) -> list:
    def _exec():
        cursor = db.aql.execute(
            query=query,
            bind_vars=bind_vars,
            batch_size=5000,
            max_runtime=300,
            count=False,
        )
        return list(cursor)
    return await run_in_threadpool(_exec)


async def _safe_query(query: str, bind_vars: dict, db, label: str) -> list:
    try:
        return await _run_export_query(query=query, bind_vars=bind_vars, db=db)
    except Exception as e:
        print(f"[export] {label} query skipped: {e}")
        return []


def _flatten_value(val):
    """Stringify nested dicts/lists so Excel cells are readable."""
    if isinstance(val, list):
        return ', '.join(str(v) for v in val if v is not None)
    if isinstance(val, dict):
        return str(val)
    return val


def _df_flatten(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        df[col] = df[col].apply(_flatten_value)
    return df


async def export_va_records_multi_sheet(
    current_user: dict,
    db: StandardDatabase,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    locations: Optional[List[str]] = None,
    date_type: Optional[str] = None,
    include_pcva: bool = False,
    include_ccva: bool = False,
    file_format: str = "excel",
) -> StreamingResponse:
    """
    Multi-sheet export:
      Sheet 1  – all form_submission variables matching the date/location filter
      Sheet 2  – PCVA results (only if include_pcva=True)
      Sheet 3  – CCVA results (only if include_ccva=True)

    VA ID appears in every sheet so the user can self-join.
    """
    try:
        print(
            f"Export: start={start_date} end={end_date} date_type={date_type} "
            f"locations={locations} include_pcva={include_pcva} "
            f"include_ccva={include_ccva} user={current_user.get('email')}"
        )

        # ── 1. Field-mapping config ──────────────────────────────────────────
        config = await fetch_odk_config(db, False)
        fm = config.field_mapping

        instance_id       = (fm and fm.instance_id)       or 'instanceid'
        region_field      = (fm and fm.location_level1)   or 'region'
        interview_date_f  = (fm and fm.interview_date)    or 'id10012'
        death_date_f      = (fm and fm.death_date)        or 'id10023'
        submission_date_f = (fm and fm.submitted_date)    or 'submissiondate'

        date_field = death_date_f
        if date_type == 'interview_date':
            date_field = interview_date_f
        elif date_type == 'submission_date':
            date_field = submission_date_f

        # ── 2. Build filter clause ───────────────────────────────────────────
        filter_conditions: list = []
        bind_vars: dict = {}

        if start_date:
            filter_conditions.append(f"va.{date_field} >= @start_date")
            bind_vars['start_date'] = start_date.isoformat()
        if end_date:
            filter_conditions.append(f"va.{date_field} <= @end_date")
            bind_vars['end_date'] = end_date.isoformat()

        if locations:
            filter_conditions.append(f"va.{region_field} IN @locations")
            bind_vars['locations'] = locations

        locationKey, locationLimitValues = get_location_limit_values(current_user)
        if locationKey and locationLimitValues:
            filter_conditions.append(f"va.{locationKey} IN @access_limit_values")
            bind_vars['access_limit_values'] = locationLimitValues

        filter_clause = " AND ".join(filter_conditions) if filter_conditions else "true"

        # ── 3. Sheet 1 – ALL form submission variables ───────────────────────
        # UNSET strips ArangoDB internals; the instance ID column is moved to front
        va_query = f"""
            FOR va IN {db_collections.VA_TABLE}
                FILTER {filter_clause}
                RETURN UNSET(va, ["_id", "_rev", "_key"])
        """
        va_records = await _safe_query(va_query, bind_vars, db, "VA records")

        if not va_records:
            raise HTTPException(
                status_code=404,
                detail="No records found matching the specified filters"
            )

        df_va = pd.DataFrame(va_records)

        # Move instance ID to first column and rename to "VA ID"
        if instance_id in df_va.columns:
            cols = [instance_id] + [c for c in df_va.columns if c != instance_id]
            df_va = df_va[cols]
        df_va = df_va.rename(columns={instance_id: 'VA ID'})
        df_va = _df_flatten(df_va)
        df_va.fillna('', inplace=True)

        va_ids = df_va['VA ID'].dropna().tolist()
        va_ids = [v for v in va_ids if v]

        print(f"Export: {len(va_records)} VA records")

        # ── 4. Sheet 2 – PCVA results (optional) ────────────────────────────
        df_pcva = None
        if include_pcva:
            pcva_records: list = []
            if db.has_collection(db_collections.PCVA_RESULTS) and va_ids:
                pcva_query = f"""
                    FOR pcva IN {db_collections.PCVA_RESULTS}
                        FILTER pcva.assigned_va IN @va_ids
                        SORT pcva.assigned_va, pcva.datetime DESC
                        COLLECT va_id = pcva.assigned_va INTO grouped = pcva
                        LET latest = FIRST(grouped)
                        RETURN {{
                            va_id:               va_id,
                            coder:               latest.created_by,
                            coded_at:            latest.datetime,
                            cause_a:             latest.frameA.a,
                            cause_b:             latest.frameA.b,
                            cause_c:             latest.frameA.c,
                            cause_d:             latest.frameA.d,
                            contributories:      latest.frameA.contributories
                        }}
                """
                pcva_records = await _safe_query(
                    pcva_query, {'va_ids': va_ids}, db, "PCVA"
                )

            # Resolve ICD10 UUIDs
            icd10_lookup: dict = {}
            if pcva_records and db.has_collection(db_collections.ICD10):
                icd10_query = f"""
                    FOR code IN {db_collections.ICD10}
                        RETURN {{ uuid: code.uuid, code: code.code, name: code.name }}
                """
                icd10_raw = await _safe_query(icd10_query, {}, db, "ICD10")
                icd10_lookup = {
                    r['uuid']: f"({r['code']}) {r['name']}"
                    for r in icd10_raw if r.get('uuid')
                }

            def _resolve(uuid):
                return icd10_lookup.get(uuid, uuid or '') if uuid else ''

            def _resolve_list(uuids):
                if not uuids:
                    return ''
                return ', '.join(filter(None, [_resolve(u) for u in uuids]))

            if pcva_records:
                for p in pcva_records:
                    p['cause_a']        = _resolve(p.get('cause_a'))
                    p['cause_b']        = _resolve(p.get('cause_b'))
                    p['cause_c']        = _resolve(p.get('cause_c'))
                    p['cause_d']        = _resolve(p.get('cause_d'))
                    underlying = next(
                        (x for x in [
                            p['cause_d'], p['cause_c'],
                            p['cause_b'], p['cause_a']
                        ] if x), ''
                    )
                    p['underlying_cause']    = underlying
                    p['contributory_causes'] = _resolve_list(p.pop('contributories', None))

                df_pcva = pd.DataFrame(pcva_records)
                df_pcva = df_pcva.rename(columns={'va_id': 'VA ID'})
                col_order = ['VA ID', 'coder', 'coded_at', 'underlying_cause',
                             'cause_a', 'cause_b', 'cause_c', 'cause_d',
                             'contributory_causes']
                df_pcva = df_pcva[[c for c in col_order if c in df_pcva.columns]]
                df_pcva = df_pcva.rename(columns={
                    'coder':              'Coder',
                    'coded_at':           'Coded At',
                    'underlying_cause':   'Underlying Cause',
                    'cause_a':            'Cause A',
                    'cause_b':            'Cause B',
                    'cause_c':            'Cause C',
                    'cause_d':            'Cause D',
                    'contributory_causes':'Contributory Causes',
                })
                df_pcva.fillna('', inplace=True)
                print(f"Export: {len(df_pcva)} PCVA records")
            else:
                df_pcva = pd.DataFrame([{'Note': 'No PCVA results found for the selected records'}])

        # ── 5. Sheet 3 – CCVA results (optional) ────────────────────────────
        df_ccva = None
        if include_ccva:
            ccva_records: list = []
            if db.has_collection(db_collections.CCVA_RESULTS) and va_ids:
                ccva_query = f"""
                    FOR ccva IN {db_collections.CCVA_RESULTS}
                        FILTER ccva.ID IN @va_ids OR ccva.uid IN @va_ids
                        FILTER ccva.CAUSE1 != null AND ccva.CAUSE1 != ""
                        COLLECT va_id = (ccva.ID != null ? ccva.ID : ccva.uid)
                        AGGREGATE
                            top_cause   = MAX(ccva.CAUSE1),
                            cause2      = MAX(ccva.CAUSE2),
                            cause3      = MAX(ccva.CAUSE3),
                            likelihood  = MAX(ccva.LIK1),
                            age_group   = MAX(ccva.age_group),
                            gender      = MAX(ccva.gender)
                        RETURN {{
                            va_id:      va_id,
                            top_cause:  top_cause,
                            cause2:     cause2,
                            cause3:     cause3,
                            likelihood: likelihood,
                            age_group:  age_group,
                            gender:     gender
                        }}
                """
                ccva_records = await _safe_query(
                    ccva_query, {'va_ids': va_ids}, db, "CCVA"
                )

            if ccva_records:
                df_ccva = pd.DataFrame(ccva_records)
                df_ccva = df_ccva.rename(columns={
                    'va_id':     'VA ID',
                    'top_cause': 'Top Cause (Cause 1)',
                    'cause2':    'Cause 2',
                    'cause3':    'Cause 3',
                    'likelihood':'Likelihood',
                    'age_group': 'Age Group',
                    'gender':    'Gender',
                })
                df_ccva.fillna('', inplace=True)
                print(f"Export: {len(df_ccva)} CCVA records")
            else:
                df_ccva = pd.DataFrame([{'Note': 'No CCVA results found for the selected records'}])

        # ── 6. Write output ──────────────────────────────────────────────────
        output = BytesIO()
        today = date.today().isoformat()

        if file_format.lower() == "csv":
            # CSV only supports one sheet — write form submissions
            df_va.index = range(1, len(df_va) + 1)
            df_va.index.name = 'No.'
            df_va.to_csv(output, index=True)
            output.seek(0)
            media_type = "text/csv"
            filename   = f"va_export_{today}.csv"
        else:
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Sheet 1
                df_va.index = range(1, len(df_va) + 1)
                df_va.index.name = 'No.'
                df_va.to_excel(writer, index=True, sheet_name="VA Data")

                # Sheet 2
                if df_pcva is not None:
                    if 'VA ID' in df_pcva.columns:
                        df_pcva.index = range(1, len(df_pcva) + 1)
                        df_pcva.index.name = 'No.'
                    df_pcva.to_excel(writer, index=True, sheet_name="PCVA Results")

                # Sheet 3
                if df_ccva is not None:
                    if 'VA ID' in df_ccva.columns:
                        df_ccva.index = range(1, len(df_ccva) + 1)
                        df_ccva.index.name = 'No.'
                    df_ccva.to_excel(writer, index=True, sheet_name="CCVA Results")

                # Strip all formatting from every sheet
                _strip_workbook_formatting(writer.book)

            output.seek(0)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename   = f"va_export_{today}.xlsx"

        headers = {"Content-Disposition": f"attachment; filename={filename}"}
        return StreamingResponse(output, media_type=media_type, headers=headers)

    except HTTPException:
        raise
    except Exception as e:
        print(f"Export error: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
